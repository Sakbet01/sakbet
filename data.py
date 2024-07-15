import tkinter as tk
from tkinter import ttk

import openpyxl
from openpyxl import Workbook, load_workbook

wb =load_workbook('records.xlsx')
ws = wb.active
ws.append(['name','Age','total_weight', 'Employment'])

wb.save('record.xlsx')

def load_file():
    path = "C:\PycharmProjects\Discount\ records.xlsx"

    workbook = openpyxl.load_workbook(path)
    sheet= workbook.active

    list_values = list(sheet.values)
    print(list_values)
    for col_name in list_values[0]:
        tree_view.heading(col_name,text=col_name)

    for value_tuple in list_values[1]:
        tree_view.insert('',tk.END, values=value_tuple)
def insert_table(name,weight,age,check):
    name = name_entry.get()
    weight = Total_weight.get()
    age = age_spinbox.get()
    check = a.get()
    check = "Employed" if a.get() else "unemployed"

roots = tk.Tk()
#style =ttk.style(root)
#root.tk.cell('source', "forest-light.tcl")
#root.tk.cell('source', "forest-dark.tcl")
#style.theme_use("forest-dark")


frame = ttk.Frame(roots)
frame.pack()


widget_frame = ttk.LabelFrame(frame, text='insert row')
widget_frame.grid(row=0,column=0 ,padx=5, pady =5)

#def insert_row():
  #  pass

name_entry = ttk.Entry(widget_frame)
name_entry.grid(row=0, column=0, sticky="nw")
name_entry.insert(0,"Name")
name_entry.bind("<FocusIn>",lambda e: name_entry.delete('0','end'))
name_entry.grid(row=0, column=0, padx=10, pady =10,sticky="nw")

Total_weight = ttk.Entry(widget_frame)
Total_weight.grid(row=1, column=0, sticky="nw")
Total_weight.insert(0,"Total weight")
Total_weight.bind("<FocusIn>",lambda e: Total_weight.delete('0','end'))
Total_weight.grid(row=1, column=0, padx=10, pady =10,sticky="ew")


age_spinbox = ttk.Spinbox(widget_frame,from_=18, to=100)
age_spinbox.bind("<FocusIn>",lambda e: age_spinbox.delete('0','end'))

age_spinbox.insert('0','age')
age_spinbox.grid(row=2, column=0)
gender= ['male','female']
status_combobox =ttk.Combobox(widget_frame, values=gender)
status_combobox.current()
status_combobox.set('male')
status_combobox.grid(row=3, column=0, padx=10, pady =10,sticky="nw")

a = tk.BooleanVar()
check_button =ttk.Checkbutton(widget_frame,text='Employed', variable=a)
check_button.grid(row=4, column=0)

button = ttk.Button(widget_frame,text = "SUBMIT", command=insert_table,)
button.grid(row =5, column =0, sticky="nsew")
separator = ttk.Separator(widget_frame)
separator.grid(row =6, column =0, padx=(50,50), pady =25, sticky="nsew")


switch_button =ttk.Checkbutton(widget_frame,text="mode",)
switch_button.grid(row=7, column=0,)

treeframe = ttk.Frame(frame)
treeframe.grid(row=0,column=1, pady=5)
treeScroll = ttk.Scrollbar(treeframe)
treeScroll.pack(side="right", fill="y")

cols =['name','Age','total_weight', 'Employment']
tree_view = ttk.Treeview(treeframe,show="headings",yscrollcommand=treeScroll.set, columns=cols,height=13)
tree_view.column("name",width=100)
tree_view.column("Age",width=50)
tree_view.column("total_weight",width=100)
tree_view.column("Employment",width=100)
tree_view.pack()
treeScroll.config(command=tree_view.yview)




roots.mainloop()